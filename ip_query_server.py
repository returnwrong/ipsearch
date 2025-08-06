from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
import ipaddress
import os

app = Flask(__name__)

class IPQuerySystem:
    def __init__(self, xlsx_file):
        self.xlsx_file = xlsx_file
        self.ip_data = self._load_ip_data()
    
    def _load_ip_data(self):
        """从XLSX文件加载IP数据"""
        wb = load_workbook(self.xlsx_file)
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 4:
                # 清理数据，去除两端空格
                ip_address = str(row[0]).strip() if row[0] else ''
                location = str(row[1]).strip() if row[1] else ''
                isp = str(row[2]).strip() if row[2] else ''
                remarks = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                
                if ip_address:  # 只有IP地址不为空才添加
                    data.append({
                        'ip_address': ip_address,
                        'location': location,
                        'isp': isp,
                        'remarks': remarks
                    })
        return data
    
    def _ip_in_range(self, ip, ip_range):
        """改进的IP范围检查方法，处理各种格式"""
        try:
            ip = ip.strip()
            ip_range = ip_range.strip()
            
            # 处理空值
            if not ip or not ip_range:
                return False
                
            # 处理CIDR格式 (192.168.1.0/24)
            if '/' in ip_range:
                network = ipaddress.ip_network(ip_range, strict=False)
                return ipaddress.ip_address(ip) in network
            
            # 处理范围格式 (192.168.1.1-192.168.1.254)
            if '-' in ip_range:
                start_ip, end_ip = ip_range.split('-')
                start_ip = start_ip.strip()
                end_ip = end_ip.strip()
                
                # 处理单IP范围情况 (如123.150.66.220-123.150.66.220)
                if start_ip == end_ip:
                    return ip == start_ip
                    
                start_ip = ipaddress.ip_address(start_ip)
                end_ip = ipaddress.ip_address(end_ip)
                ip = ipaddress.ip_address(ip)
                return start_ip <= ip <= end_ip
            
            # 精确匹配
            return ip == ip_range
            
        except ValueError as e:
            print(f"IP匹配错误: {e}, IP: {ip}, 范围: {ip_range}")
            return False
    
    def search_ip(self, ip):
        """查询IP信息"""
        print(f"正在查询IP: {ip}")  # 调试信息
        matches = []
        
        try:
            ipaddress.ip_address(ip)
        except ValueError:
            return {'success': False, 'message': '无效的IP地址格式'}
        
        for entry in self.ip_data:
            print(f"检查条目: {entry['ip_address']}")  # 调试信息
            if self._ip_in_range(ip, entry['ip_address']):
                print(f"找到匹配: {entry}")  # 调试信息
                match_type = 'exact'
                if '/' in entry['ip_address']:
                    match_type = 'network'
                elif '-' in entry['ip_address']:
                    match_type = 'range'
                
                matches.append({
                    'ip': ip,
                    'matched_network': entry['ip_address'],
                    'location': entry['location'],
                    'isp': entry['isp'],
                    'remarks': entry['remarks'],
                    'match_type': match_type
                })
        
        if matches:
            return {'success': True, 'data': matches}
        else:
            print(f"未找到匹配: {ip}")  # 调试信息
            return {'success': False, 'message': '未找到该IP地址的信息'}

# 初始化查询系统
query_system = IPQuerySystem('ip_data.xlsx')  # 替换为您的XLSX文件路径

@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

@app.route('/search', methods=['GET'])
def search():
    ip = request.args.get('ip', '').strip()
    if not ip:
        return jsonify({'success': False, 'message': '请提供IP地址'})
    return jsonify(query_system.search_ip(ip))

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('.', path)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)